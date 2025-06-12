import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook

st.title("ERP Overdue Report Generator")

uploaded_file = st.file_uploader("Upload ERP export Excel file", type=["xlsx"])

if uploaded_file:
    with st.spinner("⏳ Processing file..."):
        try:
            # Step 1: Load and preprocess input data
            df = pd.read_excel(uploaded_file)
            df['Due date'] = pd.to_datetime(df['Due date'], errors='coerce').dt.date

            df['Due date'] = pd.to_datetime(df['Due date'])
            df["Days late"] = (datetime.today() - df["Due date"]).dt.days
            filtered = df[df['Days late'] >= 2].drop(columns=["Days late"], axis=1)
            filtered = filtered[~filtered["Invoice"].astype(str).str.startswith("6")]

            # Optional: format 'Due date' for export
            filtered['Due date'] = pd.to_datetime(filtered['Due date']).dt.strftime('%Y-%m-%d')
            filtered["Date"] = pd.to_datetime(filtered['Date']).dt.strftime('%Y-%m-%d')

            # Step 2: Load template and target sheet
            wb = load_workbook("template_short.xlsx")
            sheet = wb["Open customer invoices"]

            # Step 3: Write filtered data to sheet (A–J)
            for r_idx, row in enumerate(filtered.values.tolist(), start=2):
                for c_idx, value in enumerate(row, start=1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

            # Step 4: Prepare download
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.success("✅ Report generated successfully.")
            st.download_button(
                label="⬇️ Download updated Excel file",
                data=output,
                file_name="daily_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"❌ Error during processing: {e}")
